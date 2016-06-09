from openpyxl import load_workbook
from  eTraveler.clientAPI.connection import Connection
import yaml, os, sys, string

demoStr = "_Test_999"
savedir='test'
excel_file='demo_CR_20160609.xlsx'
subsystem = 'CR'
nameIndex = 2
descIndex = 3
levelIndex = 1
maxLevel = 5
if not os.path.exists(savedir):
    os.mkdir(savedir)
print "yaml files will be saved in dir ", savedir

myConn = Connection('heather', 'Raw', prodServer=True)


class yamlAssembly():
    def __init__(self, record):
        self._yaml = yaml.load(open("assembly_CR_template.yml"))
        self._name = (record[nameIndex].value+demoStr).encode('utf-8')
        self._yaml['HardwareGroup'] = self._name
        self.registerHardwareType(self._name, record)

    def registerHardwareType(self, name, rec):
        newId = ''
        try: 
            newId = myConn.defineHardwareType(name=name, 
                                              description=rec[descIndex].value,
                                              subsystem=subsystem,
                                              batchedFlag=0,
                                              sequenceWidth='0')
            print 'New hardware type defined.  Returned id is ', newId
        except Exception,msg:
            if msg.message != "A component type with name %s already exists."%name:
                print 'Hardware type %s definition failed with exception : %s'%(name,msg.message)

    def registerRelationship(self, relname, name):
        newId = ''
        try:
            newId = myConn.defineRelationshipType(name = relname, 
                                                  description = 'rel type via eT API',
                                                  hardwareTypeName = self._name,
                                                  numItems = 1,
                                                  minorTypeName = name,
                                                  slotNames = 'na'
                                                  )
            print 'New relationship defined with id ', newId
        except Exception, msg:
            if "A relationship type with name %s"%relname not in msg.message:
                print 'Relationship definition %s failed with exception : %s'%(relname,msg.message)
        

    def save(self, ingest=False):
        filepath=os.path.join(savedir,'%s.yaml'%self._name)
        with open(filepath, 'w') as outfile:
            outfile.write( os.path.join(savedir,yaml.dump(self._yaml, default_flow_style=False)) )
            if ingest is True:
                msg = myConn.validateYaml(filepath)
                if msg==0:
                    errorCode = myConn.uploadYaml(filepath, reason="")
                    if errorCode !=0:
                        print "ingestion failed with code :", errorCode
                else:
                    "file %s validation failed with error code: %d"%(self._name, msg)
            
    def addRelationship(self, gg):
        name = (gg[nameIndex].value+demoStr).encode('utf-8') 
        if gg[levelIndex].value == maxLevel:
            self.registerHardwareType(name, gg)#the deepest elements need registration
        sequence = self._yaml['Sequence'][0]
        relname = '%s-%s'%(self._name, name)
        new_relationship = {'RelationshipName':relname, 'RelationshipAction':'install'}
        self.registerRelationship(relname, name)
        if sequence['RelationshipTasks'] is None:
            sequence['RelationshipTasks'] = [new_relationship]
        else:
            sequence['RelationshipTasks'].append(new_relationship)


try :
    print "loading the workbook ", excel_file
    wb = load_workbook(filename = excel_file, read_only=True)
    ws = wb.active
    gen = ws.iter_rows()
except:
    wb = load_workbook(filename = excel_file)
    ws = wb.active
    gen = ws.rows

print "excel sheet %s loaded"%excel_file
d0 = None
d1 = None
d2 = None
d3 = None
d4 = None
for gg in gen:
    if gg[3].value == None:
        break
    if gg[0].value != None:
        if gg[0].value == "Omit":
            continue

    if gg[1].value == 0:
#        print gg[1].value, gg[2].value
        d0 = yamlAssembly(gg)
    if gg[1].value == 1:
#        print gg[1].value, gg[2].value
        if d1 is not None:
            d1.save()
        d1 = yamlAssembly(gg)
        d0.addRelationship(gg)
    if gg[1].value == 2:
#        print gg[1].value, gg[2].value
        if d2 is not None:
            d2.save()
        d2 = yamlAssembly(gg)
        d1.addRelationship(gg)
    if gg[1].value == 3:
#        print gg[1].value, gg[2].value
        if d3 is not None:
            d3.save()
        d3 = yamlAssembly(gg)
        d2.addRelationship(gg)
    if gg[1].value == 4:
#        print gg[1].value, gg[2].value
        if d4 is not None:
            d4.save()
        d4 = yamlAssembly(gg)
        d3.addRelationship(gg)
    if gg[1].value == 5:
#        print gg[1].value, gg[2].value
        d4.addRelationship(gg)
d0.save(ingest=False)
if d1 is not None:
    d1.save()
if d2 is not None:
    d2.save()
if d3 is not None:
    d3.save()
if d4 is not None:
    d4.save()
    #     print gg[3].value 
    #     for gg in gen:
    #         if gg[0].value==2:
    #             print '\t', gg[3].value
    #             for gg in gen:
    #                 if gg[0].value==3 or gg[0].value==None:
    #                     print '\t\t', gg[3].value
    #                 else:
    #                     break
    #         else:
    #             break
